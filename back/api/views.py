from django.shortcuts import render
from django.contrib.auth.models import User
from .serializer import UserSerializer, PatrimonioSerializer
import logging
from rest_framework.decorators import api_view, parser_classes
from rest_framework import status
import json
from rest_framework.generics import ListCreateAPIView, CreateAPIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Patrimonio
from openpyxl import load_workbook

#Filters
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend

logger = logging.getLogger(__name__)

@api_view(['POST'])
@parser_classes([MultiPartParser])
def importar_patrimonios_view(request):
    try:
        if 'arquivo' not in request.FILES:
            return Response({'erro': 'Nenhum arquivo enviado.'}, status=status.HTTP_400_BAD_REQUEST)

        arquivo = request.FILES['arquivo']
        logger.info(f"Arquivo recebido: {arquivo.name}")

        try:
            dados = json.load(arquivo)
            logger.info(f"Dados do arquivo (primeiros 5): {dados[:5]}")  # Exibe apenas os primeiros 5 itens
        except json.JSONDecodeError as e:
            logger.error(f"Erro ao processar o arquivo JSON: {str(e)}")
            return Response({'erro': 'Erro ao ler o arquivo JSON. Verifique o formato.'}, status=status.HTTP_400_BAD_REQUEST)

        if not isinstance(dados, list):
            return Response({'erro': 'O arquivo JSON deve conter uma lista de objetos.'}, status=status.HTTP_400_BAD_REQUEST)

        for item in dados:
            try:
                ni = item.get('ni')
                desc = item.get('desc')
                loca = item.get('loca')

                if not all([ni, desc, loca]):
                    logger.warning(f"Faltando dados no item: {item}")
                    continue

                Patrimonio.objects.create(
                    ni=ni,
                    desc=desc,
                    loca=loca
                )
                logger.info(f"Patrimônio criado com sucesso: {item}")

            except Exception as e:
                logger.error(f"Erro ao criar patrimônio: {e}")
                continue

        return Response({'mensagem': 'Importação realizada com sucesso!'}, status=status.HTTP_201_CREATED)

    except Exception as e:
        logger.error(f"Erro desconhecido: {str(e)}")
        return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


class SignUpView(CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer


class PatrimoniosView(ListCreateAPIView):
    queryset = Patrimonio.objects.all()
    serializer_class = PatrimonioSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['ni']
    search_fields = ['desc']


class UploadXLSXView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba
        
        

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            ni = str(row[0]).strip() if row[0] is not None else None
            loca = str(row[1]).strip() if row[1] is not None else None
            desc = str(row[2]).strip() if row[2] is not None else None
            
            if not ni:
                    print(f"[Linha {i+2}] Erro: ni vazio. Dados: {row}")
                    continue  # pula a linha
                
            Patrimonio.objects.create(
                ni=row[0],
                loca=row[1],
                desc=row[2],
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})